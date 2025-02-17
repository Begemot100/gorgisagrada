from urllib.parse import quote
import math
from datetime import datetime, date, timedelta, time as dt_time
import time
from io import BytesIO
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, jsonify, session, send_file, make_response
import os
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import extract, func
import logging
from collections import defaultdict
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import hashlib
import openpyxl


logging.basicConfig(level=logging.INFO)

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', '6006')

database_url = os.getenv('DATABASE_URL')

if database_url and database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)


if not database_url:
    database_url = f"sqlite:///{os.path.join(os.getcwd(), 'instance', 'employees.db')}"

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
migrate = Migrate(app, db)

class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password, method='pbkdf2:sha256', salt_length=16)


    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class DashboardUser(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    role = db.Column(db.String(50), nullable=False)  


    def set_password(self, password):
        self.password_hash = hashlib.sha256(password.encode('utf-8')).hexdigest()

    def check_password(self, password):
        return self.password_hash == hashlib.sha256(password.encode('utf-8')).hexdigest()

class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(100), nullable=False)
    nie = db.Column(db.String(20), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=True)
    # hours_per_week = db.Column(db.Integer, nullable=False)
    days_per_week = db.Column(db.Integer, nullable=False, default=5)
    position = db.Column(db.String(50), nullable=False)
    phone = db.Column(db.String(20), nullable=False)
    email = db.Column(db.String(100), nullable=False)
    section = db.Column(db.String(50), nullable=False)
    check_in_time = db.Column(db.DateTime, nullable=True)
    check_out_time = db.Column(db.DateTime, nullable=True)
    daily_hours = db.Column(db.Float, default=0)
    monthly_hours = db.Column(db.Float, default=0)
    work_start_time = db.Column(db.Time, nullable=False)
    work_end_time = db.Column(db.Time, nullable=False)
    total_hours = db.Column(db.Float, default=0)
    total_days = db.Column(db.Integer, default=0)
    paid_holidays = db.Column(db.Integer, default=0)
    unpaid_holidays = db.Column(db.Integer, default=0)

    work_logs = db.relationship('WorkLog', backref='employee', cascade="all, delete-orphan", lazy=True)

    def __repr__(self):
        return f'<Employee {self.full_name}>'

with app.app_context():
    db.create_all()


class WorkLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    check_in_time = db.Column(db.DateTime, nullable=True)
    check_out_time = db.Column(db.DateTime, nullable=True)
    worked_hours = db.Column(db.Float, default=0)
    log_date = db.Column(db.Date, nullable=False)
    holidays = db.Column(db.String(50), default='Working day')
    # work_log = db.session.get(WorkLog, id)
    def calculate_worked_hours(self):
        if self.check_in_time and self.check_out_time:
            time_diff = self.check_out_time - self.check_in_time
            return round(time_diff.total_seconds() / 3600, 2)
        return 0

class Profile(db.Model):
    __tablename__ = 'profile'

    id = db.Column(db.Integer, primary_key=True)
    cafe_name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20), nullable=False)
    telegram = db.Column(db.String(100), nullable=True)
    instagram = db.Column(db.String(100), nullable=True)


@app.route('/')
def home():
    return render_template('home.html')


@app.route('/admin_register', methods=['GET', 'POST'])
def admin_register():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if password != confirm_password:
            return jsonify({'error': 'Passwords do not match'}), 400

        if Admin.query.filter_by(email=email).first():
            return jsonify({'error': 'Admin with this email already exists'}), 400

        admin = Admin(email=email)
        admin.set_password(password)
        db.session.add(admin)
        db.session.commit()

        return redirect('/admin_login')  

    return render_template('admin_register.html')  

@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')

        admin = Admin.query.filter_by(email=email).first()
        if not admin or not admin.check_password(password):
            return jsonify({'error': 'Invalid email or password'}), 401

        session['admin_logged_in'] = True
        session['admin_email'] = admin.email

        return redirect('/login')  

    return render_template('admin_login.html')  

@app.route('/worker_register', methods=['GET', 'POST'])
def worker_register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if password != confirm_password:
            return jsonify({'error': 'Passwords do not match'}), 400

        if DashboardUser.query.filter_by(username=username).first():
            return jsonify({'error': 'Worker with this username already exists'}), 400

        worker = DashboardUser(username=username, role='worker')
        worker.set_password(password)
        db.session.add(worker)
        db.session.commit()

        return redirect('/worker_login')  
    return render_template('worker_register.html')  

@app.route('/worker_login', methods=['GET', 'POST'])
def worker_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        worker = DashboardUser.query.filter_by(username=username, role='worker').first()
        if not worker or not worker.check_password(password):
            return jsonify({'error': 'Invalid username or password'}), 401

        session['worker_logged_in'] = True
        session['worker_username'] = worker.username

        return redirect('/board')  

    return render_template('worker_login.html')  

@app.route('/logout')
def logout():
    session.pop('admin_logged_in', None)
    session.pop('admin_email', None)
    return redirect('/admin_login')

@app.route('/worker_logout')
def worker_logout():
    session.pop('worker_logged_in', None)
    session.pop('worker_username', None)
    return redirect('/worker_login')



@app.route('/add', methods=['POST'])
def add_employee():
    nie = request.form.get('nie')

    existing_employee = Employee.query.filter_by(nie=nie).first()
    if existing_employee:
        return jsonify({'error': 'Employee with this NIE already exists'}), 400

    full_name = request.form.get('full_name')
    position = request.form.get('position')
    phone = request.form.get('phone')
    email = request.form.get('email')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    section = request.form.get('section')
    nie = request.form.get('nie')
    work_start_time = request.form.get('work_start_time')
    work_end_time = request.form.get('work_end_time')

    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None
        work_start_time = datetime.strptime(work_start_time, '%H:%M').time() if work_start_time else None
        work_end_time = datetime.strptime(work_end_time, '%H:%M').time() if work_end_time else None
    except ValueError:
        return jsonify({'error': 'Invalid date or time format'}), 400

    new_employee = Employee(
        full_name=full_name,
        nie=nie,
        position=position,
        phone=phone,
        email=email,
        start_date=start_date,
        end_date=end_date,
        section=section,
        work_start_time=work_start_time,
        work_end_time=work_end_time,
    )

    db.session.add(new_employee)
    db.session.commit()

    return jsonify({'message': 'Employee added successfully!'})


@app.route('/delete/<int:employee_id>', methods=['POST'])
def delete_employee(employee_id):
    employee = Employee.query.get(employee_id)
    if employee:
        db.session.delete(employee)
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False}), 404

@app.route('/edit/<int:id>', methods=['POST'])
def edit_employee(id):
    employee = db.session.get(Employee, id)
    if not employee:
        return jsonify({'error': 'Сотрудник не найден'}), 404

    try:
        employee.full_name = request.form.get('full_name', employee.full_name)
        employee.nie = request.form.get('nie', employee.nie)
        employee.phone = request.form.get('phone', employee.phone)
        employee.position = request.form.get('position', employee.position)
        employee.email = request.form.get('email', employee.email)
        employee.section = request.form.get('section', employee.section)
        employee.days_per_week = int(request.form.get('days_per_week', employee.days_per_week))

        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        employee.start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else employee.start_date
        employee.end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None

        work_start_time_str = request.form.get('work_start_time')
        work_end_time_str = request.form.get('work_end_time')

        if work_start_time_str:
            employee.work_start_time = datetime.strptime(work_start_time_str, '%H:%M').time()
        if work_end_time_str:
            employee.work_end_time = datetime.strptime(work_end_time_str, '%H:%M').time()

        db.session.commit()
        logging.info(f"Сотрудник {employee.full_name} успешно обновлен.")
        return jsonify({'message': 'Сотрудник успешно обновлен'}), 200

    except ValueError as e:
        logging.error(f"Ошибка при парсинге данных: {e}")
        return jsonify({'error': 'Некорректный формат данных'}), 400

    except Exception as e:
        logging.error(f"Ошибка при обновлении сотрудника: {e}")
        db.session.rollback()
        return jsonify({'error': 'Ошибка при сохранении изменений'}), 500


@app.route('/login')
def dashboard():
    employees = Employee.query.all()  
    logging.info(f"Fetched employees: {[employee.full_name for employee in employees]}")
    return render_template('index.html', employees=employees)

@app.route('/check_employees', methods=['GET'])
def check_employees():
    employees = Employee.query.all()
    return jsonify([{
        "full_name": e.full_name,
        "nie": e.nie,
        "start_date": e.start_date.isoformat() if e.start_date else None,
        "end_date": e.end_date.isoformat() if e.end_date else None,
        "work_start_time": e.work_start_time.strftime('%H:%M') if e.work_start_time else None,
        "work_end_time": e.work_end_time.strftime('%H:%M') if e.work_end_time else None,
        "days_per_week": e.days_per_week,
        "position": e.position,
        "phone": e.phone,
        "email": e.email,
        "section": e.section
    } for e in employees])

@app.route('/employee/<int:employee_id>', methods=['GET'])
def get_employee(employee_id):
    employee = db.session.get(Employee, employee_id)
    if not employee:
        return jsonify({'error': 'Employee not found'}), 404

    return jsonify({
        'id': employee.id,
        'full_name': employee.full_name,
        'nie': employee.nie,
        'start_date': employee.start_date.strftime('%Y-%m-%d') if employee.start_date else '',
        'end_date': employee.end_date.strftime('%Y-%m-%d') if employee.end_date else '',
        'work_start_time': employee.work_start_time.strftime('%H:%M') if employee.work_start_time else '',
        'work_end_time': employee.work_end_time.strftime('%H:%M') if employee.work_end_time else '',
        'days_per_week': employee.days_per_week,
        'position': employee.position,
        'section': employee.section,
        'phone': employee.phone,
        'email': employee.email
    })


@app.route('/check_in/<int:employee_id>', methods=['POST'])
def check_in(employee_id):
    try:
        current_date = date.today()

        work_log = WorkLog.query.filter_by(employee_id=employee_id, log_date=current_date).first()

        if work_log and work_log.check_in_time:
            return jsonify({"error": "Already checked in for today"}), 400

        if not work_log:
            work_log = WorkLog(
                employee_id=employee_id,
                log_date=current_date,
                check_in_time=datetime.now()
            )
            db.session.add(work_log)
        else:
            work_log.check_in_time = datetime.now()

        db.session.commit()
        logging.info(f"Check-in successful for employee_id={employee_id}, time={work_log.check_in_time}")

        return jsonify({"check_in_time": work_log.check_in_time.strftime('%H:%M')}), 200

    except Exception as e:
        logging.error(f"Error during check-in: {e}")
        return jsonify({"error": "Server error. Please try again later."}), 500

@app.route('/check_out/<int:employee_id>', methods=['POST'])
def check_out(employee_id):
    try:
        current_date = date.today()

        work_log = WorkLog.query.filter_by(employee_id=employee_id, log_date=current_date).first()

        if not work_log or not work_log.check_in_time:
            return jsonify({"error": "Cannot check out before checking in"}), 400

        if work_log.check_out_time:
            return jsonify({"error": "Already checked out for today"}), 400

        work_log.check_out_time = datetime.now()
        work_log.worked_hours = work_log.calculate_worked_hours()

        db.session.commit()
        logging.info(f"Check-out successful for employee_id={employee_id}, time={work_log.check_out_time}")

        return jsonify({
            "check_out_time": work_log.check_out_time.strftime('%H:%M'),
            "daily_hours": work_log.worked_hours
        }), 200

    except Exception as e:
        logging.error(f"Error during check-out: {e}")
        return jsonify({"error": "Server error. Please try again later."}), 500

def reset_logs():
    try:
        current_date = datetime.now().date()

        employees = Employee.query.all()

        for employee in employees:
            existing_log = WorkLog.query.filter_by(employee_id=employee.id, log_date=current_date).first()

            if not existing_log:
                new_log = WorkLog(
                    employee_id=employee.id,
                    log_date=current_date
                )
                db.session.add(new_log)

        db.session.commit()
        logging.info("Logs reset for the new day.")

    except Exception as e:
        logging.error(f"Error resetting logs: {e}")


@app.route('/reset_log/<int:log_id>', methods=['POST'])
def reset_log(log_id):
    log = WorkLog.query.get(log_id)
    if not log:
        return jsonify({'success': False, 'message': 'Work log not found'}), 404

    log.check_in_time = None
    log.check_out_time = None
    log.worked_hours = 0
    log.holidays = 'Paid'  

    try:
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


scheduler = BackgroundScheduler()
scheduler.add_job(func=reset_logs, trigger='cron', hour=0, minute=0)
scheduler.start()


def calculate_hours(check_in, check_out):
    try:
        check_in_time = datetime.strptime(check_in, '%H:%M:%S')
        check_out_time = datetime.strptime(check_out, '%H:%M:%S')
        total_seconds = (check_out_time - check_in_time).seconds
        return round(total_seconds / 3600, 2)
    except Exception as e:
        print(f"Error in calculate_hours: {e}")
        return 0.0


@app.route('/board')
def board():
    employees = Employee.query.all()
    dashboard_data = []
    today = date.today()

    for employee in employees:
        work_log = WorkLog.query.filter_by(employee_id=employee.id, log_date=today).first()

        if not work_log:
            work_log = WorkLog(
                employee_id=employee.id,
                log_date=today,
                check_in_time=None,
                check_out_time=None,
                worked_hours=0.0
            )
            db.session.add(work_log)
            db.session.commit()  

        dashboard_data.append({
            'employee': employee,
            'check_in_time': work_log.check_in_time.strftime('%H:%M') if work_log.check_in_time else '--:--',
            'check_out_time': work_log.check_out_time.strftime('%H:%M') if work_log.check_out_time else '--:--',
            'daily_hours': work_log.worked_hours if work_log.worked_hours else 0.0
        })

    return render_template('board.html', dashboard_data=dashboard_data, current_date=today)




@app.route('/work', methods=['GET'])
def work():
    group = request.args.get('group')  
    start_date_str = request.args.get('start_date')  
    end_date_str = request.args.get('end_date')  
    filter_type = request.args.get('filter', 'thismonth')  # Тип фильтра: today, yesterday, last7days и т.д.

    today = date.today()
    employees_query = Employee.query 
    if group:
        employees_query = employees_query.filter(Employee.section == group)

    employees = employees_query.all()
    employee_logs = []

    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"error": "Invalid date format"}), 400
    elif filter_type == 'today':
        start_date = end_date = today
    elif filter_type == 'yesterday':
        start_date = end_date = today - timedelta(days=1)
    elif filter_type == 'last7days':
        start_date = today - timedelta(days=7)
        end_date = today
    elif filter_type == 'last30days':
        start_date = today - timedelta(days=30)
        end_date = today
    elif filter_type == 'thismonth':
        start_date = today.replace(day=1)
        end_date = (today.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    elif filter_type == 'lastmonth':
        first_day_of_this_month = today.replace(day=1)
        last_day_of_last_month = first_day_of_this_month - timedelta(days=1)
        start_date = last_day_of_last_month.replace(day=1)
        end_date = last_day_of_last_month
    else:
        start_date = end_date = today

    for employee in employees:
        logs = WorkLog.query.filter(
            and_(
                WorkLog.employee_id == employee.id,
                WorkLog.log_date >= start_date,
                WorkLog.log_date <= end_date
            )
        ).order_by(WorkLog.log_date.desc()).all()

        paid_holidays = sum(1 for log in logs if log.holidays == 'paid')
        unpaid_holidays = sum(1 for log in logs if log.holidays == 'unpaid')

        employee_logs.append({
            'employee': {
                'id': employee.id,
                'full_name': employee.full_name,
                'position': employee.position,
                'section': employee.section  
            },
            'work_logs': [{
                'id': log.id,
                'log_date': log.log_date.strftime('%Y-%m-%d'),
                'check_in_time': log.check_in_time.strftime('%H:%M') if log.check_in_time else '--:--',
                'check_out_time': log.check_out_time.strftime('%H:%M') if log.check_out_time else '--:--',
                'worked_hours': log.worked_hours or 0,
                'holidays': log.holidays
            } for log in logs],
            'total_hours': round(sum(log.worked_hours or 0 for log in logs), 2),
            'total_days': len(logs),
            'paid_holidays': paid_holidays,
            'unpaid_holidays': unpaid_holidays,
        })

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_template('partials/work_logs.html', employees=employee_logs)
    else:
        return render_template('work.html', employees=employee_logs, today=today)
    # return render_template("work.html", employees=employee_logs, today=today)


@app.route('/update_holiday_status/<int:log_id>', methods=['POST'])
def update_holiday_status(log_id):
    app.logger.info(f"Получен запрос на обновление статуса для log_id={log_id}")

    data = request.get_json()
    status = data.get('status')
    if not status:
        return jsonify({'success': False, 'message': 'Status not provided'}), 400

    log = WorkLog.query.get(log_id)
    if not log:
        return jsonify({'success': False, 'message': 'Log not found'}), 404

    employee = Employee.query.get(log.employee_id)
    if not employee:
        return jsonify({'success': False, 'message': 'Employee not found'}), 404

    log.holidays = status
    if status in ["paid", "unpaid"]:
        log.check_in_time = None
        log.check_out_time = None
        log.worked_hours = 0

    if status == 'unpaid' and employee.total_days > 0:
        employee.total_days -= 1

    try:
        db.session.commit()
        return jsonify({'success': True, 'status': log.holidays})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.template_filter('format_hours')
def format_hours(value):
    if value is None:
        return '0h 0min'
    hours = int(value)
    minutes = int((value - hours) * 60)
    return f'{hours}h {minutes}min'


@app.route('/update_log_time/<int:log_id>', methods=['POST'])
def update_log_time(log_id):
    data = request.get_json()
    check_in_time = data.get('check_in_time')  
    check_out_time = data.get('check_out_time')  
    if not check_in_time or not check_out_time:
        return jsonify({"success": False, "message": "Invalid input"}), 400

    try:
        check_in_time_obj = datetime.strptime(check_in_time, "%H:%M")
        check_out_time_obj = datetime.strptime(check_out_time, "%H:%M")
    except ValueError:
        return jsonify({"success": False, "message": "Invalid time format"}), 400

    work_log = WorkLog.query.get(log_id)
    if not work_log:
        return jsonify({"success": False, "message": "Work log not found"}), 404

    work_log.check_in_time = check_in_time_obj
    work_log.check_out_time = check_out_time_obj

    worked_hours = (check_out_time_obj - check_in_time_obj).total_seconds() / 3600.0
    work_log.worked_hours = round(worked_hours, 2)

    db.session.commit()

    return jsonify({"success": True, "message": "Work log updated successfully"})



def add_missing_logs():
    employees = Employee.query.all()
    today = date.today()

    for employee in employees:
        existing_log = WorkLog.query.filter_by(employee_id=employee.id, log_date=today).first()
        if not existing_log:
            new_log = WorkLog(
                employee_id=employee.id,
                log_date=today,
                check_in_time=None,
                check_out_time=None,
                worked_hours=0
            )
            db.session.add(new_log)

    db.session.commit()



@app.route('/export_excel', methods=['POST'])
def export_excel():
    selected_employees = request.json.get('selectedEmployees', [])

    if not selected_employees:
        return {"message": "No employees selected"}, 400

    # Fetch data for selected employees
    employees_data = []
    for employee_id in selected_employees:
        employee = Employee.query.get(employee_id)
        if employee:
            logs = WorkLog.query.filter_by(employee_id=employee_id).all()
            employees_data.append({
                "employee": employee,
                "logs": logs
            })

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Logs"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")
    row = 1
    for data in employees_data:
        employee = data["employee"]
        logs = data["logs"]

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f"{employee.full_name} - {employee.position} {employee.nie}")
        ws.cell(row=row, column=1).font = bold_font
        row += 1

        headers = ["Date", "Entrada", "Salida", "Total Day Hours", "Holiday Type", "Days Worked"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = bold_font
            cell.alignment = center_alignment
        row += 1

        for log in logs:
            ws.cell(row=row, column=1, value=log.log_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=log.check_in_time or '--:--')
            ws.cell(row=row, column=3, value=log.check_out_time or '--:--')
            ws.cell(row=row, column=4, value=log.worked_hours or "0:00")
            ws.cell(row=row, column=5, value=log.holidays.capitalize() if log.holidays else "Working day")
            ws.cell(row=row, column=6, value=1 if log.worked_hours else 0)
            row += 1

        total_days = sum(1 for log in logs if log.worked_hours)
        total_hours = sum(log.worked_hours for log in logs if log.worked_hours)
        ws.cell(row=row, column=5, value="Total Worked Days")
        ws.cell(row=row, column=6, value=total_days)
        row += 1

        ws.cell(row=row, column=5, value="Total Hours")
        ws.cell(row=row, column=6, value=f"{total_hours} hours")
        row += 3  # Space between employees (2 lines + 1 extra increment)

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2  

    file_path = os.path.join("instance", "employee_logs.xlsx")
    wb.save(file_path)

    return send_file(file_path, as_attachment=True, download_name="employee_logs.xlsx")

@app.route('/export_employees_excel', methods=['POST'])
def export_employees_excel():
    # Get employee IDs from the request
    employee_ids = request.json.get('employeeIds', [])

    if not employee_ids:
        return {"message": "No employees selected for export."}, 400

    employees = Employee.query.filter(Employee.id.in_(employee_ids)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Employees"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")

    headers = ["ID", "Name", "Position", "NIE", "Phone", "Email", "Start Date", "Start Time", "End Time"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold_font
        cell.alignment = center_alignment

    for row, employee in enumerate(employees, start=2):
        ws.cell(row=row, column=1, value=employee.id)
        ws.cell(row=row, column=2, value=employee.full_name)
        ws.cell(row=row, column=3, value=employee.position)
        ws.cell(row=row, column=4, value=employee.nie)
        ws.cell(row=row, column=5, value=employee.phone)
        ws.cell(row=row, column=6, value=employee.email)
        ws.cell(row=row, column=7, value=employee.start_date.strftime('%Y-%m-%d') if employee.start_date else '')
        ws.cell(row=row, column=8, value=employee.work_start_time.strftime('%H:%M') if employee.work_start_time else '')
        ws.cell(row=row, column=9, value=employee.work_end_time.strftime('%H:%M') if employee.work_end_time else '')

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    file_path = os.path.join("instance", "filtered_employees.xlsx")
    wb.save(file_path)

    return send_file(file_path, as_attachment=True, download_name="filtered_employees.xlsx")


@app.route('/filter_by_group', methods=['POST'])
def filter_by_group():
    data = request.get_json()
    groups = data.get('groups', [])

    employees = Employee.query.filter(Employee.section.in_(groups)).all()
    result = [
        {
            'id': employee.id,
            'full_name': employee.full_name,
            'position': employee.position,
            'group': employee.section
        }
        for employee in employees
    ]

    return jsonify(result)



@app.route('/update_profile', methods=['POST'])
def update_profile():
    try:
        cafe_name = request.form.get('cafe_name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        telegram = request.form.get('telegram')
        instagram = request.form.get('instagram')

        if not cafe_name or not email or not phone:
            return jsonify({"success": False, "message": "All required fields must be filled."})

        profile = Profile.query.first()  
        if profile:
            profile.cafe_name = cafe_name
            profile.email = email
            profile.phone = phone
            profile.telegram = telegram
            profile.instagram = instagram
            db.session.commit()
        else:
            return jsonify({"success": False, "message": "Profile not found."})

        return jsonify({"success": True, "message": "Profile updated successfully!"})
    except Exception as e:
        print(f"Error updating profile: {e}")
        return jsonify({"success": False, "message": "An error occurred while updating the profile."})


scheduler = BackgroundScheduler()
scheduler.add_job(func=add_missing_logs, trigger="cron", hour=23, minute=45)  
scheduler.start()


if __name__ == '__main__':

    app.run(debug=False, port=5002)
